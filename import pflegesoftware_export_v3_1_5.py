# Erweiterungen für v3.1.5 Market-Ready
# Teil 1: Excel/CSV-Export für Pflegesoftware-Import

import pandas as pd
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

# ============================================================================
# PFLEGESOFTWARE-EXPORT ENGINE
# ============================================================================

class PflegesoftwareExporter:
    """
    Generiert strukturierte Exporte für:
    - DM7 Connext (CSV)
    - Vivendi PD (Excel)
    - Medifox DAN (Excel)
    
    Ermöglicht Bulk-Import in Pflegesoftware
    """
    
    # DM7 Connext Spalten-Mapping
    DM7_COLUMNS = [
        'PatientenID',
        'Name',
        'Vorname',
        'Geburtsdatum',
        'Aufnahmedatum',
        'Aufgenommen_durch',
        'Pflegegrad',
        'RIA_Sturzrisiko',
        'RIA_Dekubitus',
        'RIA_Ernaehrung',
        'RIA_Medikamente',
        'FEM_vorhanden',
        'FEM_Typ',
        'FEM_Beschluss',
        'DVA_Compliance',
        'Compliance_Score',
        'MDK_Ready',
        'Themenfeld_1_Mobilität',
        'Themenfeld_2_Kognition',
        'Themenfeld_3_Krankheit',
        'Themenfeld_4_Selbstversorgung',
        'Themenfeld_5_Alltagsleben',
        'Themenfeld_6_Soziales',
        'Bemerkungen'
    ]
    
    # Vivendi PD Mapping
    VIVENDI_COLUMNS = [
        'Bewohner_ID',
        'Nachname',
        'Vorname',
        'Geb_Datum',
        'Einzug',
        'Pflegegrad',
        'NBA_Modul_1',
        'NBA_Modul_2',
        'NBA_Modul_3',
        'NBA_Modul_4',
        'NBA_Modul_5',
        'NBA_Modul_6',
        'Risiko_Sturz',
        'Risiko_Dekubitus',
        'Risiko_Mangelernährung',
        'FEM_Maßnahmen',
        'Dokumentation_vollständig',
        'Notizen'
    ]
    
    # Medifox DAN Mapping
    MEDIFOX_COLUMNS = [
        'ID',
        'Name',
        'Aufnahme',
        'PG',
        'Risiken',
        'Maßnahmen',
        'FEM',
        'Status'
    ]
    
    @classmethod
    def export_dm7_csv(cls, assessments: List['AssessmentResult'], filepath: str) -> str:
        """
        Exportiert Assessments als DM7-kompatible CSV
        
        Args:
            assessments: Liste von AssessmentResult-Objekten
            filepath: Pfad zur Output-Datei
            
        Returns:
            Erfolgsmeldung
        """
        data = []
        
        for assessment in assessments:
            # Patientendaten extrahieren
            row = {
                'PatientenID': assessment.patient_id,
                'Name': assessment.patient_name.split()[0] if ' ' in assessment.patient_name else assessment.patient_name,
                'Vorname': ' '.join(assessment.patient_name.split()[1:]) if ' ' in assessment.patient_name else '',
                'Geburtsdatum': '',  # Nicht im Assessment
                'Aufnahmedatum': assessment.aufnahme_datum.strftime('%d.%m.%Y'),
                'Aufgenommen_durch': assessment.aufgenommen_durch,
                'Pflegegrad': '',  # Wird aus BI-Modulen berechnet
                
                # RIA-Trigger
                'RIA_Sturzrisiko': cls._has_ria_trigger(assessment, 'Sturzrisiko'),
                'RIA_Dekubitus': cls._has_ria_trigger(assessment, 'Dekubitus'),
                'RIA_Ernaehrung': cls._has_ria_trigger(assessment, 'Ernährung'),
                'RIA_Medikamente': cls._has_ria_trigger(assessment, 'Medikamente'),
                
                # FEM
                'FEM_vorhanden': 'Ja' if assessment.fem_alerts else 'Nein',
                'FEM_Typ': ', '.join([alert.detected_keyword for alert in assessment.fem_alerts]) if assessment.fem_alerts else '',
                'FEM_Beschluss': 'Ja' if any(alert.beschluss_vorhanden for alert in assessment.fem_alerts) else 'Nein',
                
                # DVA & Compliance
                'DVA_Compliance': assessment.overall_compliance.value,
                'Compliance_Score': f"{assessment.compliance_score:.0f}%",
                'MDK_Ready': 'Ja' if assessment.mdk_ready else 'Nein',
                
                # SiS-Themenfelder
                'Themenfeld_1_Mobilität': cls._get_sis_field(assessment, 'Mobilität'),
                'Themenfeld_2_Kognition': cls._get_sis_field(assessment, 'Kognition'),
                'Themenfeld_3_Krankheit': cls._get_sis_field(assessment, 'Krankheit'),
                'Themenfeld_4_Selbstversorgung': cls._get_sis_field(assessment, 'Selbstversorgung'),
                'Themenfeld_5_Alltagsleben': cls._get_sis_field(assessment, 'Alltagsleben'),
                'Themenfeld_6_Soziales': cls._get_sis_field(assessment, 'Soziales'),
                
                'Bemerkungen': ''
            }
            
            data.append(row)
        
        # Als CSV speichern (Semikolon-getrennt, UTF-8 mit BOM für Excel)
        df = pd.DataFrame(data, columns=cls.DM7_COLUMNS)
        df.to_csv(filepath, sep=';', index=False, encoding='utf-8-sig')
        
        return f"✅ DM7-CSV erfolgreich erstellt: {filepath}"
    
    @classmethod
    def export_vivendi_excel(cls, assessments: List['AssessmentResult'], filepath: str) -> str:
        """
        Exportiert Assessments als Vivendi PD Excel
        
        Args:
            assessments: Liste von AssessmentResult-Objekten
            filepath: Pfad zur Output-Datei
            
        Returns:
            Erfolgsmeldung
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Aufnahmen"
        
        # Header-Zeile (fett, blau)
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(cls.VIVENDI_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Daten-Zeilen
        for row_num, assessment in enumerate(assessments, 2):
            ws.cell(row=row_num, column=1, value=assessment.patient_id)
            
            # Name splitten
            name_parts = assessment.patient_name.split()
            ws.cell(row=row_num, column=2, value=name_parts[0] if name_parts else '')
            ws.cell(row=row_num, column=3, value=' '.join(name_parts[1:]) if len(name_parts) > 1 else '')
            
            ws.cell(row=row_num, column=4, value='')  # Geburtsdatum
            ws.cell(row=row_num, column=5, value=assessment.aufnahme_datum.strftime('%d.%m.%Y'))
            ws.cell(row=row_num, column=6, value='')  # Pflegegrad
            
            # NBA-Module
            for i in range(6):
                module_data = cls._get_bi_module(assessment, i+1)
                ws.cell(row=row_num, column=7+i, value=module_data)
            
            # Risiken
            ws.cell(row=row_num, column=13, value=cls._has_ria_trigger(assessment, 'Sturzrisiko'))
            ws.cell(row=row_num, column=14, value=cls._has_ria_trigger(assessment, 'Dekubitus'))
            ws.cell(row=row_num, column=15, value=cls._has_ria_trigger(assessment, 'Ernährung'))
            
            # FEM
            fem_text = ', '.join([alert.detected_keyword for alert in assessment.fem_alerts]) if assessment.fem_alerts else 'Keine'
            ws.cell(row=row_num, column=16, value=fem_text)
            
            # Dokumentation
            ws.cell(row=row_num, column=17, value='Ja' if assessment.mdk_ready else 'Nein')
            ws.cell(row=row_num, column=18, value='')  # Notizen
        
        # Spaltenbreite anpassen
        for col_num in range(1, len(cls.VIVENDI_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        wb.save(filepath)
        return f"✅ Vivendi-Excel erfolgreich erstellt: {filepath}"
    
    @classmethod
    def export_medifox_excel(cls, assessments: List['AssessmentResult'], filepath: str) -> str:
        """
        Exportiert Assessments als Medifox DAN Excel
        
        Args:
            assessments: Liste von AssessmentResult-Objekten
            filepath: Pfad zur Output-Datei
            
        Returns:
            Erfolgsmeldung
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Medifox Import"
        
        # Medifox-spezifische Formatierung (grün)
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(cls.MEDIFOX_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Daten (kompakt für Medifox)
        for row_num, assessment in enumerate(assessments, 2):
            ws.cell(row=row_num, column=1, value=assessment.patient_id)
            ws.cell(row=row_num, column=2, value=assessment.patient_name)
            ws.cell(row=row_num, column=3, value=assessment.aufnahme_datum.strftime('%d.%m.%Y'))
            ws.cell(row=row_num, column=4, value='')  # PG
            
            # Risiken (kompakt)
            risiken = []
            if cls._has_ria_trigger(assessment, 'Sturzrisiko') == 'Ja':
                risiken.append('Sturz')
            if cls._has_ria_trigger(assessment, 'Dekubitus') == 'Ja':
                risiken.append('Dekubitus')
            if cls._has_ria_trigger(assessment, 'Ernährung') == 'Ja':
                risiken.append('Ernährung')
            ws.cell(row=row_num, column=5, value=', '.join(risiken) if risiken else 'Keine')
            
            # Maßnahmen
            ws.cell(row=row_num, column=6, value='Siehe Pflegeplanung')
            
            # FEM
            fem_status = 'Ja' if assessment.fem_alerts else 'Nein'
            ws.cell(row=row_num, column=7, value=fem_status)
            
            # Status
            status = 'MDK-ready' if assessment.mdk_ready else 'Prüfung erforderlich'
            ws.cell(row=row_num, column=8, value=status)
        
        # Spaltenbreite
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 20
        
        wb.save(filepath)
        return f"✅ Medifox-Excel erfolgreich erstellt: {filepath}"
    
    @classmethod
    def _has_ria_trigger(cls, assessment: 'AssessmentResult', trigger_name: str) -> str:
        """Prüft ob RIA-Trigger vorhanden"""
        for trigger in assessment.ria_triggers:
            if trigger_name.lower() in trigger.name.lower():
                return 'Ja'
        return 'Nein'
    
    @classmethod
    def _get_sis_field(cls, assessment: 'AssessmentResult', field_name: str) -> str:
        """Holt SiS-Themenfeld"""
        for sis in assessment.sis_strukturen:
            if field_name.lower() in sis.themenfeld.lower():
                return sis.information[:100] + '...' if len(sis.information) > 100 else sis.information
        return ''
    
    @classmethod
    def _get_bi_module(cls, assessment: 'AssessmentResult', module_id: int) -> str:
        """Holt BI-Modul-Daten"""
        for module in assessment.bi_modules:
            if module.module_id == module_id:
                return f"{module.points}/{module.max_points}"
        return ''

# ============================================================================
# INTEGRATION IN STREAMLIT UI
# ============================================================================

# In main() nach Smart-Copy-Sektion hinzufügen:

def show_pflegesoftware_export_ui(result):
    """
    Zeigt UI für Pflegesoftware-Export
    """
    st.markdown("---")
    st.markdown("## 📊 Export für Pflegesoftware")
    
    st.info("""
    **Bulk-Import in Ihre Pflegesoftware:**
    - DM7 Connext: CSV-Datei direkt importieren
    - Vivendi PD: Excel-Datei in Bewohner-Verwaltung laden
    - Medifox DAN: Excel-Datei kopieren & einfügen
    """)
    
    col1, col2, col3 = st.columns(3)
    
    # DM7 CSV
    with col1:
        if st.button("📥 DM7 CSV", use_container_width=True, type="secondary"):
            if 'assessments_history' in st.session_state and st.session_state['assessments_history']:
                filepath = f"/tmp/dm7_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                message = PflegesoftwareExporter.export_dm7_csv(
                    st.session_state['assessments_history'],
                    filepath
                )
                
                with open(filepath, 'rb') as f:
                    st.download_button(
                        label="💾 DM7-CSV herunterladen",
                        data=f,
                        file_name=f"dm7_import_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                st.success(message)
            else:
                st.warning("Keine Assessments vorhanden")
    
    # Vivendi Excel
    with col2:
        if st.button("📥 Vivendi Excel", use_container_width=True, type="secondary"):
            if 'assessments_history' in st.session_state and st.session_state['assessments_history']:
                filepath = f"/tmp/vivendi_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                message = PflegesoftwareExporter.export_vivendi_excel(
                    st.session_state['assessments_history'],
                    filepath
                )
                
                with open(filepath, 'rb') as f:
                    st.download_button(
                        label="💾 Vivendi-Excel herunterladen",
                        data=f,
                        file_name=f"vivendi_import_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                st.success(message)
            else:
                st.warning("Keine Assessments vorhanden")
    
    # Medifox Excel
    with col3:
        if st.button("📥 Medifox Excel", use_container_width=True, type="secondary"):
            if 'assessments_history' in st.session_state and st.session_state['assessments_history']:
                filepath = f"/tmp/medifox_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                message = PflegesoftwareExporter.export_medifox_excel(
                    st.session_state['assessments_history'],
                    filepath
                )
                
                with open(filepath, 'rb') as f:
                    st.download_button(
                        label="💾 Medifox-Excel herunterladen",
                        data=f,
                        file_name=f"medifox_import_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                st.success(message)
            else:
                st.warning("Keine Assessments vorhanden")
    
    # Import-Anleitungen
    with st.expander("📖 Import-Anleitungen"):
        st.markdown("""
        ### DM7 Connext
        1. CSV-Datei herunterladen
        2. DM7 öffnen → Bewohner → Import
        3. CSV-Datei auswählen
        4. Spalten-Mapping prüfen (automatisch)
        5. Importieren klicken
        
        ### Vivendi PD
        1. Excel-Datei herunterladen
        2. Vivendi → Bewohner-Verwaltung
        3. Excel-Import-Funktion
        4. Datei auswählen
        5. Import starten
        
        ### Medifox DAN
        1. Excel-Datei herunterladen
        2. In Excel öffnen
        3. Daten kopieren (Strg+A, Strg+C)
        4. In Medifox einfügen (Strg+V)
        5. Speichern
        """)
